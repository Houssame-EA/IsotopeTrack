"""Excel export for the Clustering Analysis results.

Writes one workbook with a sheet per kind of result rather than a single
nested document, because the things people do with these numbers are
table-shaped: put the cluster characterisation in a paper, sort the evaluation
sweep to see which K won, check how the samples split across clusters.

Sheet map
---------
``Summary``       Configuration, optimal K and algorithm, dataset size.
``Evaluation``    One row per (algorithm, K) with every metric as a column.
``Clusters``      One row per cluster: type, particle count, dominant elements.
``Composition``   One row per (cluster, element): mean, median, std, frequency.
``Samples``       One row per (cluster, sample): count and fraction.
``Stability``     Bootstrap Jaccard per cluster, mean particle stability.
``Particles``     Per-particle stability and GMM membership, when computed.

A workbook is not a faithful round-trip — floats are written at display
precision and the nesting is flattened — so the JSON export is kept alongside
it for anyone who needs to reload the exact values.
"""

from __future__ import annotations

import datetime
import logging

import numpy as np

_log = logging.getLogger("IsotopeTrack.results.cluster.export_workbook")

_HEADER_FILL = 'FF2563EB'
_HEADER_FONT = 'FFFFFFFF'
_MAX_ROWS = 1_000_000
_MAX_CHARS = 32_000
_INLINE_ITEMS = 24


def _cell(v):
    """Coerce any value into something openpyxl can write.

    Excel accepts strings, numbers, booleans, dates and None. A NumPy array or
    an arbitrary object raises ``ValueError`` deep inside openpyxl, so every
    value passes through here rather than trusting each sheet builder — the
    configuration in particular can hold whole per-particle arrays.

    Long sequences are summarised rather than dumped: a cell holding 23,411
    sample names helps nobody, and Excel caps a cell at 32,767 characters.

    Args:
        v: Any value.

    Returns:
        A string, number, bool or None.
    """
    if v is None or isinstance(v, (str, bool, int, float)):
        if isinstance(v, float) and not np.isfinite(v):
            return None
        return v[:_MAX_CHARS] if isinstance(v, str) else v
    if isinstance(v, np.generic):
        return _cell(v.item())
    if isinstance(v, np.ndarray):
        if v.size <= _INLINE_ITEMS:
            return _cell(', '.join(str(x) for x in v.tolist()))
        return '<%d values>' % v.size
    if isinstance(v, dict):
        if len(v) <= _INLINE_ITEMS:
            return _cell('; '.join('%s=%s' % (a, b) for a, b in sorted(
                v.items(), key=lambda kv: str(kv[0]))))
        return '<%d entries>' % len(v)
    if isinstance(v, (list, tuple, set, frozenset)):
        seq = list(v)
        if len(seq) <= _INLINE_ITEMS:
            return _cell(', '.join(str(x) for x in seq))
        return '<%d values>' % len(seq)
    if isinstance(v, (datetime.datetime, datetime.date, datetime.time)):
        return v
    return _cell(str(v))


def _pairs(seq):
    """Normalise an element list into ``(element, percentage)`` pairs.

    ``composition`` and ``dominant_elements`` hold pairs, but a bare element
    name is tolerated so that one malformed entry cannot cost the whole sheet.

    Args:
        seq: A sequence of pairs or of element names.

    Returns:
        list[tuple]: Pairs of element and percentage, the latter possibly None.
    """
    out = []
    for item in _as_array(seq).tolist() if isinstance(
            seq, np.ndarray) else list(seq or []):
        if isinstance(item, (list, tuple)) and len(item) >= 2:
            out.append((item[0], item[1]))
        elif isinstance(item, (list, tuple)) and len(item) == 1:
            out.append((item[0], None))
        else:
            out.append((item, None))
    return out


def _autosize(ws):
    """Widen every column to fit its longest cell.

    Args:
        ws (Worksheet): The sheet to adjust.
    """
    from openpyxl.utils import get_column_letter
    widths = {}
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is None:
                continue
            widths[cell.column] = max(widths.get(cell.column, 0),
                                      len(str(cell.value)))
    for col, w in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = min(48, w + 3)


def _write_sheet(wb, title, headers, rows):
    """Add one sheet with a styled header row and frozen panes.

    Args:
        wb (Workbook): The workbook to add to.
        title (str): Sheet name.
        headers (list[str]): Column headings.
        rows (list[list]): Row values.

    Returns:
        Worksheet | None: The sheet, or None when there was nothing to write.
    """
    from openpyxl.styles import Alignment, Font, PatternFill
    if not rows:
        return None
    ws = wb.create_sheet(title=title[:31])
    ws.append(list(headers))
    fill = PatternFill('solid', fgColor=_HEADER_FILL)
    font = Font(bold=True, color=_HEADER_FONT)
    for cell in ws[1]:
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(vertical='center')
    for r in rows[:_MAX_ROWS]:
        ws.append([_cell(v) for v in r])
    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = ws.dimensions
    _autosize(ws)
    return ws


def _f(v, nd=4):
    """Round a value for display, passing non-numbers through.

    Args:
        v: The value.
        nd (int): Decimal places.

    Returns:
        The rounded float, or the original value when it is not numeric.
    """
    try:
        if v is None:
            return None
        f = float(v)
        return None if not np.isfinite(f) else round(f, nd)
    except (TypeError, ValueError):
        return v


def _summary_rows(dlg):
    """Build the Summary sheet rows.

    Args:
        dlg: The Clustering Analysis dialog.

    Returns:
        list[list]: ``[setting, value]`` pairs.
    """
    cfg = dict(getattr(dlg.node, 'config', {}) or {})
    rows = [
        ['Optimal K', getattr(dlg, 'optimal_k', None)],
        ['Optimal algorithm', getattr(dlg, 'optimal_algorithm', None)
         or getattr(dlg, 'optimal_algo', None)],
        ['', ''],
        ['— Configuration —', ''],
    ]
    for k in sorted(cfg, key=str):
        rows.append([k, cfg[k]])
    return rows


def _evaluation_rows(dlg):
    """Build the Evaluation sheet.

    Args:
        dlg: The Clustering Analysis dialog.

    Returns:
        tuple[list[str], list[list]]: Headers and rows.
    """
    results = getattr(dlg, 'eval_results', {}) or {}
    metric_keys = []
    for res in results.values():
        for key in res:
            if key != 'k_values' and key not in metric_keys:
                metric_keys.append(key)

    try:
        from results.cluster.dialog import METRIC_REGISTRY
    except Exception:
        METRIC_REGISTRY = {}
    label = {spec['key']: spec.get('display', spec['key'])
             for spec in METRIC_REGISTRY.values() if 'key' in spec}

    headers = ['Algorithm', 'K'] + [label.get(k, k) for k in metric_keys]
    rows = []
    for algo, res in results.items():
        ks = res.get('k_values') or []
        for i, k in enumerate(ks):
            row = [algo, k]
            for key in metric_keys:
                seq = res.get(key) or []
                row.append(_f(seq[i]) if i < len(seq) else None)
            rows.append(row)
    return headers, rows


def _cluster_rows(dlg):
    """Build the Clusters sheet.

    Args:
        dlg: The Clustering Analysis dialog.

    Returns:
        tuple[list[str], list[list]]: Headers and rows.
    """
    try:
        from results.cluster.dialog import _cluster_label_short
    except Exception:
        def _cluster_label_short(cid):
            """Fall back to a plain cluster name.

            Args:
                cid (int): Cluster id.

            Returns:
                str: The label.
            """
            return 'Noise' if cid < 0 else 'C%d' % (int(cid) + 1)

    headers = ['Algorithm', 'Cluster', 'Type', 'Particles', 'Share (%)',
               'Dominant elements', 'Top composition (%)']
    rows = []
    for algo, clusters in (getattr(dlg, 'characterisation', {}) or {}).items():
        total = sum(cd.get('particle_count', 0) for cd in clusters.values())
        for cid, cd in sorted(clusters.items(),
                              key=lambda kv: -kv[1].get('particle_count', 0)):
            dom = cd.get('dominant_elements') or []
            comp = cd.get('composition') or []
            n = cd.get('particle_count', 0)
            rows.append([
                algo,
                _cluster_label_short(cid),
                cd.get('cluster_type') or '',
                n,
                _f(n / total * 100, 2) if total else None,
                ' · '.join(str(e) for e, _ in _pairs(dom)),
                ' · '.join('%s %s' % (e, _f(p, 1))
                           for e, p in _pairs(comp)[:5]),
            ])
    return headers, rows


def _composition_rows(dlg):
    """Build the Composition sheet.

    Args:
        dlg: The Clustering Analysis dialog.

    Returns:
        tuple[list[str], list[list]]: Headers and rows.
    """
    try:
        from results.cluster.dialog import _cluster_label_short
    except Exception:
        def _cluster_label_short(cid):
            """Fall back to a plain cluster name.

            Args:
                cid (int): Cluster id.

            Returns:
                str: The label.
            """
            return 'Noise' if cid < 0 else 'C%d' % (int(cid) + 1)

    headers = ['Algorithm', 'Cluster', 'Element', 'Share (%)', 'Mean',
               'Median', 'Std dev', 'Detected in', 'Detection frequency']
    rows = []
    for algo, clusters in (getattr(dlg, 'characterisation', {}) or {}).items():
        for cid, cd in sorted(clusters.items()):
            pcts = cd.get('element_pcts') or {}
            stats = cd.get('element_stats') or {}
            for el in sorted(stats, key=lambda e: -pcts.get(e, 0)):
                st = stats[el] or {}
                rows.append([
                    algo,
                    _cluster_label_short(cid),
                    el,
                    _f(pcts.get(el), 3),
                    _f(st.get('mean')),
                    _f(st.get('median')),
                    _f(st.get('std')),
                    st.get('count'),
                    _f(st.get('frequency')),
                ])
    return headers, rows


def _sample_rows(dlg):
    """Build the Samples sheet.

    Args:
        dlg: The Clustering Analysis dialog.

    Returns:
        tuple[list[str], list[list]]: Headers and rows.
    """
    try:
        from results.cluster.dialog import _cluster_label_short
    except Exception:
        def _cluster_label_short(cid):
            """Fall back to a plain cluster name.

            Args:
                cid (int): Cluster id.

            Returns:
                str: The label.
            """
            return 'Noise' if cid < 0 else 'C%d' % (int(cid) + 1)

    headers = ['Algorithm', 'Cluster', 'Sample', 'Particles', 'Fraction']
    rows = []
    for algo, clusters in (getattr(dlg, 'characterisation', {}) or {}).items():
        for cid, cd in sorted(clusters.items()):
            for sname, sd in sorted((cd.get('sample_breakdown') or {}).items()):
                rows.append([algo, _cluster_label_short(cid), sname,
                             sd.get('count'), _f(sd.get('fraction'))])
    return headers, rows


def _as_array(v):
    """Return ``v`` as a NumPy array, treating None as empty.

    ``v or []`` cannot be used here: for a NumPy array that calls ``bool()``
    on it, which raises for anything with more than one element.

    Args:
        v: An array, a sequence, or None.

    Returns:
        numpy.ndarray: The values, empty when there were none.
    """
    if v is None:
        return np.asarray([])
    return np.asarray(v)


def _stability_rows(dlg):
    """Build the Stability sheet.

    Args:
        dlg: The Clustering Analysis dialog.

    Returns:
        tuple[list[str], list[list]]: Headers and rows.
    """
    cs = getattr(dlg, 'cluster_stability', None)
    if not cs:
        return [], []
    try:
        from results.cluster.dialog import _cluster_label_short
    except Exception:
        def _cluster_label_short(cid):
            """Fall back to a plain cluster name.

            Args:
                cid (int): Cluster id.

            Returns:
                str: The label.
            """
            return 'Noise' if int(cid) < 0 else 'C%d' % (int(cid) + 1)

    headers = ['Metric', 'Cluster', 'Value']
    rows = [['Algorithm', '', cs.get('algo')],
            ['Bootstrap replicates', '', cs.get('n_boot')]]
    ps = _as_array(cs.get('particle_stability'))
    if ps.size:
        rows.append(['Mean particle stability', '', _f(ps.mean())])
    for cid, v in sorted((cs.get('cluster_jaccard') or {}).items(),
                         key=lambda kv: str(kv[0])):
        rows.append(['Cluster Jaccard', _cluster_label_short(int(cid)), _f(v)])
    return headers, rows


def _particle_rows(dlg):
    """Build the Particles sheet from the per-particle arrays.

    Only written when a bootstrap or a Gaussian Mixture fit produced
    per-particle numbers; otherwise the sheet is skipped entirely.

    Args:
        dlg: The Clustering Analysis dialog.

    Returns:
        tuple[list[str], list[list]]: Headers and rows.
    """
    cs = getattr(dlg, 'cluster_stability', None) or {}
    stab = _as_array(cs.get('particle_stability'))
    gmm = (getattr(dlg, 'final_results', {}) or {}).get('Gaussian Mixture', {})
    maxp = _as_array(gmm.get('max_proba'))

    n = max(stab.size, maxp.size)
    if not n:
        return [], []

    headers = ['Particle']
    if stab.size:
        headers.append('Assignment stability')
    if maxp.size:
        headers.append('GMM max membership')

    rows = []
    for i in range(n):
        row = [i]
        if stab.size:
            row.append(_f(stab[i]) if i < stab.size else None)
        if maxp.size:
            row.append(_f(maxp[i]) if i < maxp.size else None)
        rows.append(row)
    return headers, rows


def _algo_blocks(ws):
    """Find the contiguous row range each algorithm occupies.

    The Evaluation and Clusters sheets list every algorithm one after another
    in column A, so a chart that treated the column as a single series would
    draw a line joining the last K of one algorithm to the first K of the next.

    Args:
        ws (Worksheet): A sheet whose first column holds the algorithm name.

    Returns:
        list[tuple]: ``(name, first_row, last_row)`` per block.
    """
    blocks = []
    for r in range(2, ws.max_row + 1):
        name = ws.cell(row=r, column=1).value
        if blocks and blocks[-1][0] == name:
            blocks[-1][2] = r
        else:
            blocks.append([name, r, r])
    return [tuple(b) for b in blocks]


def _chart_evaluation(wb, ws):
    """Add one line chart per metric to the Evaluation sheet.

    Each metric gets its own chart rather than sharing one: the scales are
    incomparable — silhouette runs 0 to 1 while Calinski-Harabasz runs into the
    thousands — so a shared axis would flatten every curve but the largest.
    Within a chart each algorithm is a separate series, which is what makes the
    elbow and the chosen K readable.

    Args:
        wb (Workbook): The workbook, unused but kept for symmetry.
        ws (Worksheet): The Evaluation sheet.
    """
    from openpyxl.chart import LineChart, Reference
    from openpyxl.chart.data_source import AxDataSource, NumRef
    from openpyxl.chart.series_factory import SeriesFactory

    blocks = _algo_blocks(ws)
    if not blocks or ws.max_row < 3:
        return
    anchor = ws.max_row + 3
    for col in range(3, ws.max_column + 1):
        title = ws.cell(row=1, column=col).value
        ch = LineChart()
        ch.title = '%s vs K' % title
        ch.y_axis.title = str(title)
        ch.x_axis.title = 'K'
        ch.height, ch.width = 7.5, 14
        ch.style = 2
        for name, first, last in blocks:
            if last <= first:
                continue
            sr = SeriesFactory(
                Reference(ws, min_col=col, min_row=first, max_row=last),
                title=str(name))
            sr.cat = AxDataSource(numRef=NumRef(f=str(
                Reference(ws, min_col=2, min_row=first, max_row=last))))
            sr.smooth = False
            sr.marker.symbol = 'circle'
            sr.marker.size = 6
            ch.series.append(sr)
        if not ch.series:
            continue
        if len(ch.series) == 1:
            ch.legend = None
        ws.add_chart(ch, 'A%d' % anchor)
        anchor += 16


def _chart_clusters(wb, ws):
    """Add a bar chart of particle counts and a pie chart of the shares.

    The bar chart labels categories by algorithm and cluster together, since
    every algorithm restarts its numbering at C1. The pie covers only the first
    algorithm — shares summed across algorithms would each count the same
    particles again and total well over 100%.

    Args:
        wb (Workbook): The workbook, unused but kept for symmetry.
        ws (Worksheet): The Clusters sheet.
    """
    from openpyxl.chart import BarChart, PieChart, Reference

    blocks = _algo_blocks(ws)
    n = ws.max_row
    if n < 2 or not blocks:
        return

    bar = BarChart()
    bar.type = 'col'
    bar.title = 'Particles per cluster'
    bar.y_axis.title = 'Particles'
    bar.x_axis.title = 'Cluster'
    bar.height, bar.width = 8, 15
    bar.style = 10
    bar.add_data(Reference(ws, min_col=4, min_row=1, max_row=n),
                 titles_from_data=True)
    cat_min = 1 if len(blocks) > 1 else 2
    bar.set_categories(Reference(ws, min_col=cat_min, max_col=2, min_row=2,
                                 max_row=n))
    bar.legend = None
    ws.add_chart(bar, 'A%d' % (n + 3))

    name, first, last = blocks[0]
    if last > first:
        pie = PieChart()
        pie.title = 'Share of particles — %s' % name
        pie.height, pie.width = 8, 11
        pie.add_data(Reference(ws, min_col=5, min_row=first, max_row=last))
        pie.set_categories(Reference(ws, min_col=2, min_row=first,
                                     max_row=last))
        ws.add_chart(pie, 'J%d' % (n + 3))


def _chart_samples(wb, ws):
    """Add a stacked bar chart of how samples split across the clusters.

    Args:
        wb (Workbook): The workbook, unused but kept for symmetry.
        ws (Worksheet): The Samples sheet.
    """
    from openpyxl.chart import BarChart, Reference

    n = ws.max_row
    if n < 2:
        return
    ch = BarChart()
    ch.type = 'col'
    ch.grouping = 'stacked'
    ch.overlap = 100
    ch.title = 'Particles by cluster and sample'
    ch.y_axis.title = 'Particles'
    ch.x_axis.title = 'Cluster · Sample'
    ch.height, ch.width = 8, 16
    ch.style = 12
    ch.add_data(Reference(ws, min_col=4, min_row=1, max_row=n),
                titles_from_data=True)
    ch.set_categories(Reference(ws, min_col=2, max_col=3, min_row=2, max_row=n))
    ch.legend = None
    ws.add_chart(ch, 'A%d' % (n + 3))


def _chart_stability(wb, ws):
    """Add a bar chart of the per-cluster bootstrap Jaccard scores.

    Args:
        wb (Workbook): The workbook, unused but kept for symmetry.
        ws (Worksheet): The Stability sheet.
    """
    from openpyxl.chart import BarChart, Reference

    rows = [r for r in range(2, ws.max_row + 1)
            if ws.cell(row=r, column=1).value == 'Cluster Jaccard']
    if len(rows) < 2:
        return
    first, last = rows[0], rows[-1]
    ch = BarChart()
    ch.type = 'col'
    ch.title = 'Cluster stability (Jaccard)'
    ch.y_axis.title = 'Jaccard'
    ch.y_axis.scaling.min = 0
    ch.y_axis.scaling.max = 1
    ch.x_axis.title = 'Cluster'
    ch.height, ch.width = 8, 14
    ch.style = 10
    ch.add_data(Reference(ws, min_col=3, min_row=first, max_row=last))
    ch.set_categories(Reference(ws, min_col=2, min_row=first, max_row=last))
    ch.legend = None
    ws.add_chart(ch, 'E2')


def export_workbook(dlg, path):
    """Write the clustering results to an Excel workbook.

    Args:
        dlg: The Clustering Analysis dialog holding the results.
        path (str): Destination ``.xlsx`` path.

    Returns:
        list[str]: The names of the sheets written.

    Raises:
        ImportError: When openpyxl is unavailable.
    """
    from openpyxl import Workbook

    wb = Workbook()
    wb.remove(wb.active)

    _write_sheet(wb, 'Summary', ['Setting', 'Value'], _summary_rows(dlg))
    for title, builder, charter in (
            ('Evaluation', _evaluation_rows, _chart_evaluation),
            ('Clusters', _cluster_rows, _chart_clusters),
            ('Composition', _composition_rows, None),
            ('Samples', _sample_rows, _chart_samples),
            ('Stability', _stability_rows, _chart_stability),
            ('Particles', _particle_rows, None)):
        try:
            headers, rows = builder(dlg)
        except Exception:
            _log.exception("could not build the %s sheet", title)
            continue
        ws = _write_sheet(wb, title, headers, rows)
        if ws is None or charter is None:
            continue
        try:
            charter(wb, ws)
        except Exception:
            _log.exception("could not chart the %s sheet", title)

    if not wb.sheetnames:
        _write_sheet(wb, 'Summary', ['Setting', 'Value'],
                     [['No results', 'Run the clustering first']])
    wb.save(path)
    return list(wb.sheetnames)
