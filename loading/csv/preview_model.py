"""Lazy, chunk-backed table model for the file-import preview.

The preview exposes a small window of rows at first and pulls further rows from
disk only when the view scrolls towards the bottom, so opening a multi-gigabyte
time series costs no more than opening a hundred-row file. There is no ceiling
on how far the user can scroll: the model keeps fetching until the file ends.

The model also renders the keep/remove state owned by ``ExclusionManager``.
Removed rows and columns stay visible but are dimmed and struck through rather
than vanishing, so a removal is always something the user can see and reverse.
"""
from __future__ import annotations

import logging
from pathlib import Path

import numpy as np
import pandas as pd
from PySide6.QtCore import QAbstractTableModel, QModelIndex, Qt, Signal
from PySide6.QtGui import QBrush, QColor, QFont

_itk_log = logging.getLogger("IsotopeTrack.loading.csv.preview_model")

DISK_CHUNK_ROWS = 2000
INITIAL_VISIBLE_ROWS = 20
FETCH_STEP_ROWS = 100
INITIAL_VISIBLE_COLUMNS = 20
FETCH_STEP_COLUMNS = 20
MIN_GRID_ROWS = 20
MIN_GRID_COLUMNS = 20

DELIMITED_EXTS = {'.csv', '.txt'}
EXCEL_EXTS = {'.xls', '.xlsx', '.xlsm', '.xlsb'}

CANDIDATE_DELIMITERS = [',', ';', '\t', '|']
CANDIDATE_ENCODINGS = ['utf-8-sig', 'utf-16', 'cp1252', 'latin-1']


def detect_encoding(path: str | Path) -> str:
    """Return an encoding that can decode the start of a file.

    Byte-order marks are checked first because they are definitive, then a
    short sample is trial-decoded. Latin-1 is last and always succeeds, so the
    function never fails to return something usable.

    Args:
        path (str | Path): File to inspect.

    Returns:
        str: Name of an encoding that decodes the file's opening bytes.
    """
    try:
        with open(path, 'rb') as handle:
            raw = handle.read(65536)
    except OSError:
        return 'utf-8-sig'

    if raw.startswith(b'\xef\xbb\xbf'):
        return 'utf-8-sig'
    if raw.startswith((b'\xff\xfe', b'\xfe\xff')):
        return 'utf-16'

    for encoding in CANDIDATE_ENCODINGS:
        try:
            raw.decode(encoding)
        except (UnicodeDecodeError, LookupError):
            continue
        return encoding
    return 'latin-1'


LAYOUT_SCAN_LINES = 120
LAYOUT_SCAN_LIMIT = 20000


def _sample_lines(path: str | Path, encoding: str,
                  limit: int = LAYOUT_SCAN_LINES) -> list[str]:
    """Return the first non-empty lines of a file, with line endings stripped.

    Args:
        path (str | Path): File to read.
        encoding (str): Encoding to decode with.
        limit (int): Maximum number of lines to return.

    Returns:
        list[str]: The sampled lines, or an empty list if the file is unreadable.
    """
    try:
        with open(path, 'r', encoding=encoding, errors='replace') as handle:
            lines = []
            while len(lines) < limit:
                raw = handle.readline()
                if not raw:
                    break
                stripped = raw.rstrip('\r\n')
                if stripped.strip():
                    lines.append(stripped)
            return lines
    except OSError:
        return []


def _numeric_shape(fields) -> tuple[int, int]:
    """Return how many of a split line's values are numbers.

    Args:
        fields: The line already split on a candidate separator.

    Returns:
        tuple[int, int]: Count of values that parse as numbers, and count of
            non-empty values.
    """
    filled = [f.strip() for f in fields if f.strip()]
    if not filled:
        return 0, 0
    parsed = 0
    for value in filled:
        try:
            float(value)
        except ValueError:
            continue
        parsed += 1
    return parsed, len(filled)


def _longest_data_run(rows, minimum_values: int) -> tuple[int, int, int]:
    """Return the longest stretch of consecutive rows that look like readings.

    A row counts as data when most of its values are numbers. Half is enough
    rather than nearly all, because runs commonly carry a label or flag column
    alongside the measurements, and demanding that every value be numeric would
    reject the very rows being looked for.

    Args:
        rows: Lines already split on a candidate separator.
        minimum_values (int): How many numbers a row needs before it can count
            as data.

    Returns:
        tuple[int, int, int]: Length of the run, its width, and where it starts.
    """
    flags = []
    widths = []
    for fields in rows:
        numeric, filled = _numeric_shape(fields)
        flags.append(numeric >= minimum_values and numeric * 2 >= filled)
        widths.append(filled)

    best = (0, 0, 0)
    index = 0
    while index < len(flags):
        if not flags[index]:
            index += 1
            continue
        end = index
        while end + 1 < len(flags) and flags[end + 1]:
            end += 1
        candidate = (end - index + 1, widths[index], index)
        if candidate[:2] > best[:2]:
            best = candidate
        index = end + 1
    return best


def detect_layout(path: str | Path, encoding: str) -> tuple[str, int]:
    """Return the separator and the line that names the columns.

    Instrument exports open with a block of metadata before the real header: an
    acquisition path, an operator name, a batch date. Those lines are found by
    looking for where the *numbers* begin rather than by counting fields,
    because exports routinely pad every line to the same width with trailing
    separators. A file whose metadata lines end in ``,,,`` has exactly as many
    fields as its data rows, and counting alone would call line one the header.

    The data is the longest run of consecutive lines whose values parse as
    numbers; the header is the line immediately above it.

    A short sample is tried first and widened only if it found nothing, so the
    common file costs a hundred lines of reading while a file with an unusually
    long preamble is still found rather than silently mis-read.

    Args:
        path (str | Path): File to inspect.
        encoding (str): Encoding to read the sample with.

    Returns:
        tuple[str, int]: The separator, and how many lines precede the header.
    """
    for limit in (LAYOUT_SCAN_LINES, LAYOUT_SCAN_LIMIT):
        lines = _sample_lines(path, encoding, limit)
        if not lines:
            return ',', 0

        for minimum in (2, 1):
            best_delimiter = ','
            best_run = (0, 0, 0)
            for candidate in CANDIDATE_DELIMITERS:
                rows = [line.split(candidate) for line in lines]
                run = _longest_data_run(rows, minimum)
                if run[:2] > best_run[:2]:
                    best_run = run
                    best_delimiter = candidate
            if best_run[0]:
                return best_delimiter, max(0, best_run[2] - 1)

        if len(lines) < limit:
            break

    return ',', 0


def detect_delimiter(path: str | Path, encoding: str, skip_rows: int = 0) -> str:
    """Return the field separator a delimited file appears to use.

    Args:
        path (str | Path): File to inspect.
        encoding (str): Encoding to read the sample with.
        skip_rows (int): Accepted for call compatibility and ignored.

    Returns:
        str: The detected separator, defaulting to a comma.
    """
    return detect_layout(path, encoding)[0]


def sniff_delimited_settings(path: str | Path, skip_rows: int | None = None) -> dict:
    """Return the parse settings a delimited file appears to need.

    Args:
        path (str | Path): File to inspect.
        skip_rows (int | None): Force a number of leading lines to skip instead
            of detecting it, for when the user overrides the guess.

    Returns:
        dict: Detected ``encoding``, ``delimiter`` and ``skip_rows``.
    """
    encoding = detect_encoding(path)
    delimiter, detected_skip = detect_layout(path, encoding)
    width, full_width = detect_table_width(path, encoding, delimiter)
    return {
        'encoding': encoding,
        'delimiter': delimiter,
        'skip_rows': detected_skip if skip_rows is None else max(0, skip_rows),
        'width': width,
        'full_width': full_width,
    }


def detect_table_width(path: str | Path, encoding: str,
                       delimiter: str) -> tuple[int, int]:
    """Return how many columns a file has, before and after trimming padding.

    Exports often end every line with a run of separators, which parses as
    columns that are empty from top to bottom. They are an artefact of the
    writer, not data, so they are trimmed from what the user is shown. The
    untrimmed count is still needed to parse the lines correctly.

    Args:
        path (str | Path): File to inspect.
        encoding (str): Encoding to read the sample with.
        delimiter (str): Separator to split on.

    Returns:
        tuple[int, int]: Columns worth showing, and columns actually present.
    """
    lines = _sample_lines(path, encoding, LAYOUT_SCAN_LIMIT)
    if not lines:
        return 1, 1

    rows = [line.split(delimiter) for line in lines]
    full = max(1, max(len(row) for row in rows))

    width = full
    while width > 1:
        column = width - 1
        if any(len(row) > column and row[column].strip() for row in rows):
            break
        width -= 1
    return width, full


def describe_delimiter(delimiter: str) -> str:
    """Return a readable name for a separator character.

    Args:
        delimiter (str): The separator.

    Returns:
        str: A word such as ``"comma"`` or ``"tab"``.
    """
    return {
        ',': 'comma',
        ';': 'semicolon',
        '\t': 'tab',
        '\\t': 'tab',
        '|': 'pipe',
        ' ': 'space',
    }.get(delimiter, f"'{delimiter}'")


def file_type_of(path: str | Path) -> str:
    """Return ``'delimited'``, ``'excel'`` or ``'unknown'`` for a file path.

    Args:
        path (str | Path): File or directory path.

    Returns:
        str: The coarse file family used to pick a reader.
    """
    ext = Path(path).suffix.lower()
    if ext in DELIMITED_EXTS:
        return 'delimited'
    if ext in EXCEL_EXTS:
        return 'excel'
    return 'unknown'


NUMERIC_COLUMN_RATIO = 0.7


def numeric_like_columns(df: pd.DataFrame) -> list:
    """Return the columns that carry measurements rather than labels.

    A column counts as numeric-like when it is already a numeric dtype, or when
    most of its non-empty values parse as numbers. The second case matters
    because a single stray word anywhere in a column makes pandas read the
    whole column as text.

    Args:
        df (pd.DataFrame): Frame to inspect.

    Returns:
        list: Column labels holding numeric data.
    """
    numeric: list = []
    for column in df.columns:
        series = df[column]
        if pd.api.types.is_numeric_dtype(series):
            numeric.append(column)
            continue
        text = series.astype(str).str.strip()
        filled = text[(text != '') & (text.str.lower() != 'nan')]
        if filled.empty:
            continue
        parsed = pd.to_numeric(filled, errors='coerce')
        if parsed.notna().mean() >= NUMERIC_COLUMN_RATIO:
            numeric.append(column)
    return numeric


def find_first_stopping_row(df: pd.DataFrame) -> int:
    """Return the index of the first row where the usable data ends.

    Instrument exports often append a footer such as "End of acquisition" after
    the last reading. Such a row is recognised by its measurement columns
    failing to hold a number, not by the mere presence of text: a file with a
    genuine label column, like a per-row comment or a flag, has text in every
    row and must not be cut at row zero.

    A row that is entirely empty also ends the data, which covers the blank
    separator line many exports place before their footer.

    Args:
        df (pd.DataFrame): Frame to scan.

    Returns:
        int: Row index of the first stopping row, or ``len(df)`` if there is none.
    """
    if df.empty:
        return 0

    blank = _blank_mask(df)
    row_all_blank = blank.all(axis=1).to_numpy()

    numeric = numeric_like_columns(df)
    if not numeric:
        if not row_all_blank.any():
            return len(df)
        return int(np.argmax(row_all_blank))

    unusable = np.ones(len(df), dtype=bool)
    for column in numeric:
        series = df[column]
        if not pd.api.types.is_numeric_dtype(series):
            series = pd.to_numeric(series.astype(str).str.strip(),
                                   errors='coerce')
        unusable &= series.isna().to_numpy()

    bad = unusable | row_all_blank
    if not bad.any():
        return len(df)
    return int(np.argmax(bad))


def _blank_mask(df: pd.DataFrame) -> pd.DataFrame:
    """Return a boolean frame marking cells that hold nothing usable.

    Args:
        df (pd.DataFrame): Frame to inspect.

    Returns:
        pd.DataFrame: True where a cell is empty, whitespace or the text "nan".
    """
    text = df.astype(str).apply(lambda s: s.str.strip().str.lower())
    return df.isna() | (text == '') | (text == 'nan')


class RowSource:
    """Base class for incremental row providers used by the preview model."""

    def __init__(self, path: str | Path):
        """Record the source path and initialise the exhaustion flags.

        Args:
            path (str | Path): File the source reads from.
        """
        self.path = str(path)
        self.columns: list[str] = []
        self._done = False
        self._pending: pd.DataFrame | None = None
        self._truncated_at: int | None = None
        self._rows_served = 0
        self.scan_from = 0

    def fetch(self, count: int) -> pd.DataFrame:
        """Return up to ``count`` further rows.

        Args:
            count (int): Maximum number of rows wanted.

        Returns:
            pd.DataFrame: The next rows, or an empty frame at end of data.
        """
        raise NotImplementedError

    def is_exhausted(self) -> bool:
        """Return True once the file has no further usable rows to hand out."""
        return self._done and not self._has_pending()

    def truncated_at(self) -> int | None:
        """Return the row where trailing text stopped the read, if it did."""
        return self._truncated_at

    def close(self) -> None:
        """Release any handle held by the source and mark it exhausted."""
        self._done = True
        self._pending = None

    def _has_pending(self) -> bool:
        """Return True when buffered rows are still waiting to be served."""
        return self._pending is not None and len(self._pending) > 0

    def _empty(self) -> pd.DataFrame:
        """Return an empty frame carrying the source's column names."""
        return pd.DataFrame(columns=self.columns)

    def _assemble(self, count: int) -> pd.DataFrame:
        """Pull chunks until ``count`` rows are gathered or the file ends.

        Args:
            count (int): Maximum number of rows wanted.

        Returns:
            pd.DataFrame: The gathered rows with a fresh positional index.
        """
        frames: list[pd.DataFrame] = []
        gathered = 0
        while gathered < count:
            chunk = self._next_chunk()
            if chunk is None or not len(chunk):
                self._done = True
                break
            stop = self._stopping_row(chunk, self._rows_served + gathered)
            if stop < len(chunk):
                self._truncated_at = self._rows_served + gathered + stop
                chunk = chunk.iloc[:stop]
                self._done = True
                self._pending = None
                if len(chunk):
                    frames.append(chunk)
                    gathered += len(chunk)
                break
            frames.append(chunk)
            gathered += len(chunk)

        if not frames:
            return self._empty()

        frame = pd.concat(frames, ignore_index=True)
        if len(frame) > count:
            self._pending = frame.iloc[count:].reset_index(drop=True)
            frame = frame.iloc[:count].reset_index(drop=True)
        self._rows_served += len(frame)
        return frame

    def _stopping_row(self, chunk: pd.DataFrame, offset: int) -> int:
        """Return where a chunk's usable data ends, ignoring any preamble.

        In raw mode the metadata lines above the header have no numbers in
        them, so scanning from row zero would end the data before it started.
        ``scan_from`` marks the first row that belongs to the table.

        Args:
            chunk (pd.DataFrame): Rows just read.
            offset (int): Absolute index of the chunk's first row.

        Returns:
            int: Index within the chunk where the data ends.
        """
        skip = max(0, self.scan_from - offset)
        if skip >= len(chunk):
            return len(chunk)
        stop = find_first_stopping_row(chunk.iloc[skip:])
        if stop >= len(chunk) - skip:
            return len(chunk)
        return stop + skip

    def _next_chunk(self) -> pd.DataFrame | None:
        """Return the next raw chunk from the buffer or the reader.

        Buffered rows are served before the exhausted flag is honoured. A read
        that overshoots the requested count leaves the surplus buffered, and if
        the same read also found the end of the data the flag would otherwise
        make those buffered rows unreachable while still reporting the source
        as having more to give, which spins the caller forever.
        """
        if self._has_pending():
            chunk = self._pending
            self._pending = None
            return chunk
        if self._done:
            return None
        return self._read_chunk()

    def _read_chunk(self) -> pd.DataFrame | None:
        """Return the next chunk straight from the underlying reader."""
        raise NotImplementedError


class DelimitedRowSource(RowSource):
    """Incremental reader for CSV and TXT files backed by a pandas iterator."""

    def __init__(self, path: str | Path, delimiter: str = ",",
                 encoding: str = "utf-8", skip_rows: int = 0,
                 header_row: int | None = 0, width: int | None = None,
                 full_width: int | None = None):
        """Open ``path`` for chunked reading and capture its column names.

        Passing ``header_row=None`` reads the file raw: every physical line
        becomes a row, including any metadata above the table, and the columns
        are positional. That is what lets the preview show a file exactly as it
        sits on disk so the user can point at the header themselves.

        Args:
            path (str | Path): File to read.
            delimiter (str): Field separator, with ``"\\t"`` accepted as a tab.
            encoding (str): Text encoding; ``utf-8`` is upgraded to ``utf-8-sig``
                so a byte-order mark never corrupts the first column name.
            skip_rows (int): Leading rows to discard before the header.
            header_row (int | None): Header row index, or None to read raw.
            width (int | None): Columns to keep when reading raw.
            full_width (int | None): Columns actually present in the file, which
                can exceed ``width`` when the export pads its lines.
        """
        super().__init__(path)
        self.delimiter = "\t" if delimiter == "\\t" else (delimiter or ",")
        self.encoding = self._normalise_encoding(encoding)
        self.skip_rows = max(0, int(skip_rows))
        self.header_row = header_row
        self.width = width
        self.full_width = full_width or width
        self._iterator = None
        self._open()

    @staticmethod
    def _normalise_encoding(encoding: str) -> str:
        """Return an encoding name that tolerates a leading byte-order mark.

        Args:
            encoding (str): Encoding chosen by the user.

        Returns:
            str: The encoding to hand to pandas.
        """
        name = (encoding or "utf-8").strip().lower()
        if name in ("utf-8", "utf8"):
            return "utf-8-sig"
        return encoding or "utf-8"

    def _read_kwargs(self, **extra) -> dict:
        """Return the shared ``read_csv`` arguments merged with ``extra``.

        Args:
            **extra: Additional keyword arguments for ``pandas.read_csv``.

        Returns:
            dict: Keyword arguments describing this source.
        """
        kwargs: dict = {
            'delimiter': self.delimiter,
            'encoding': self.encoding,
            'on_bad_lines': 'warn',
        }
        if self.header_row is not None and self.header_row >= 0:
            kwargs['header'] = self.header_row
            if self.width and self.full_width and self.full_width > self.width:
                kwargs['usecols'] = list(range(self.width))
        else:
            kwargs['header'] = None
            if self.width:
                kwargs['names'] = list(range(self.full_width or self.width))
                if self.full_width and self.full_width > self.width:
                    kwargs['usecols'] = list(range(self.width))
        if self.skip_rows > 0:
            kwargs['skiprows'] = list(range(self.skip_rows))
        kwargs.update(extra)
        return kwargs

    def _open(self) -> None:
        """Read the header and create the chunk iterator, retrying on encoding errors."""
        if self.header_row is None:
            self.columns = [str(i) for i in range(self.width or 1)]
        else:
            try:
                head = pd.read_csv(self.path, **self._read_kwargs(nrows=0))
            except UnicodeDecodeError:
                _itk_log.warning(
                    "Encoding %s failed for %s; retrying with replacement characters",
                    self.encoding, self.path)
                self.encoding = 'utf-8'
                head = pd.read_csv(
                    self.path,
                    **self._read_kwargs(nrows=0, encoding_errors='replace'))
            self.columns = [str(c) for c in head.columns]

        extra: dict = {'chunksize': DISK_CHUNK_ROWS}
        if self.encoding == 'utf-8':
            extra['encoding_errors'] = 'replace'
        self._iterator = pd.read_csv(self.path, **self._read_kwargs(**extra))

    def _read_chunk(self) -> pd.DataFrame | None:
        """Return the next chunk from the pandas iterator, or None at the end."""
        if self._iterator is None:
            return None
        try:
            chunk = next(self._iterator)
        except StopIteration:
            return None
        except Exception:
            _itk_log.exception("Chunk read failed for %s", self.path)
            return None
        chunk.columns = self.columns[:len(chunk.columns)]
        return chunk

    def fetch(self, count: int) -> pd.DataFrame:
        """Return up to ``count`` further rows from the file.

        Args:
            count (int): Maximum number of rows wanted.

        Returns:
            pd.DataFrame: The next rows, or an empty frame at end of data.
        """
        if self._done and not self._has_pending():
            return self._empty()
        return self._assemble(count)

    def close(self) -> None:
        """Close the pandas iterator and mark the source exhausted."""
        try:
            if self._iterator is not None:
                self._iterator.close()
        except Exception:
            _itk_log.debug("Iterator already closed for %s", self.path)
        self._iterator = None
        super().close()


class ExcelRowSource(RowSource):
    """Incremental reader for Excel workbooks backed by an openpyxl row iterator."""

    def __init__(self, path: str | Path, sheet_index: int = 0,
                 skip_rows: int = 0, header_row: int | None = 0):
        """Open ``path`` in read-only mode and capture its column names.

        Args:
            path (str | Path): Workbook to read.
            sheet_index (int): Zero-based index of the sheet to read.
            skip_rows (int): Leading rows to discard before the header.
            header_row (int | None): Header row index, or None for no header.
        """
        super().__init__(path)
        self.sheet_index = max(0, int(sheet_index))
        self.skip_rows = max(0, int(skip_rows))
        self.header_row = header_row
        self._workbook = None
        self._rows = None
        self._width = 0
        self._open()

    def _open(self) -> None:
        """Load the workbook, skip leading rows and read the header row."""
        try:
            import openpyxl
        except ImportError as exc:
            raise ImportError(
                "openpyxl is required for Excel files. "
                "Install with: pip install openpyxl") from exc

        self._workbook = openpyxl.load_workbook(
            self.path, read_only=True, data_only=True)
        names = self._workbook.sheetnames
        index = min(self.sheet_index, max(0, len(names) - 1))
        sheet = self._workbook[names[index]]
        self._rows = sheet.iter_rows(values_only=True)

        for _ in range(self.skip_rows):
            if next(self._rows, None) is None:
                self._done = True
                return

        if self.header_row is not None and self.header_row >= 0:
            header = next(self._rows, None)
            if header is None:
                self._done = True
                self.columns = []
                return
            self.columns = [
                str(v) if v is not None else f"Column {i + 1}"
                for i, v in enumerate(header)
            ]
        else:
            probe = next(self._rows, None)
            if probe is None:
                self._done = True
                self.columns = []
                return
            self.columns = [f"Column {i + 1}" for i in range(len(probe))]
            self._pending = self._frame_from_rows([probe])
        self._width = len(self.columns)

    def _frame_from_rows(self, rows: list[tuple]) -> pd.DataFrame:
        """Build a frame from raw openpyxl tuples, padding short rows.

        Args:
            rows (list[tuple]): Raw cell-value tuples straight from openpyxl.

        Returns:
            pd.DataFrame: Frame with this source's column names.
        """
        width = self._width or len(self.columns)
        shaped = [
            (list(r) + [None] * width)[:width] if r is not None else [None] * width
            for r in rows
        ]
        frame = pd.DataFrame(shaped, columns=self.columns[:width])
        return frame.infer_objects()

    def _read_chunk(self) -> pd.DataFrame | None:
        """Return the next block of worksheet rows, or None at the end."""
        if self._rows is None:
            return None
        batch: list[tuple] = []
        for _ in range(DISK_CHUNK_ROWS):
            row = next(self._rows, None)
            if row is None:
                break
            batch.append(row)
        if not batch:
            return None
        return self._frame_from_rows(batch)

    def fetch(self, count: int) -> pd.DataFrame:
        """Return up to ``count`` further rows from the worksheet.

        Args:
            count (int): Maximum number of rows wanted.

        Returns:
            pd.DataFrame: The next rows, or an empty frame at end of data.
        """
        if self._done and not self._has_pending():
            return self._empty()
        return self._assemble(count)

    def close(self) -> None:
        """Close the workbook and mark the source exhausted."""
        try:
            if self._workbook is not None:
                self._workbook.close()
        except Exception:
            _itk_log.debug("Workbook already closed for %s", self.path)
        self._workbook = None
        self._rows = None
        super().close()


def build_row_source(path: str | Path, settings: dict,
                     raw: bool = False) -> RowSource:
    """Return the right ``RowSource`` for ``path`` given the import settings.

    Args:
        path (str | Path): File to read.
        settings (dict): Import settings holding delimiter, encoding, sheet
            index, skipped rows and header row.
        raw (bool): True to read every physical line, preamble included, with
            positional columns. Footer detection then starts at the header row
            rather than at the top of the file.

    Returns:
        RowSource: A reader positioned at the first row it should serve.

    Raises:
        ValueError: If the file extension is not a supported import format.
    """
    kind = file_type_of(path)
    header_row = None if raw else settings.get('header_row', 0)
    skip_rows = 0 if raw else settings.get('skip_rows', 0)

    if kind == 'delimited':
        source = DelimitedRowSource(
            path,
            delimiter=settings.get('delimiter', ','),
            encoding=settings.get('encoding', 'utf-8'),
            skip_rows=skip_rows,
            header_row=header_row,
            width=settings.get('width'),
            full_width=settings.get('full_width'),
        )
    elif kind == 'excel':
        source = ExcelRowSource(
            path,
            sheet_index=settings.get('sheet_index', settings.get('sheet_name', 0)) or 0,
            skip_rows=skip_rows,
            header_row=header_row,
        )
    else:
        raise ValueError(f"Unsupported file format: {Path(path).suffix}")

    if raw:
        source.scan_from = settings.get('skip_rows', 0) + 1
    return source


def read_columns_only(path: str | Path, settings: dict) -> list[str]:
    """Return a file's column names without reading any data rows.

    Args:
        path (str | Path): File to inspect.
        settings (dict): Import settings describing how to parse the header.

    Returns:
        list[str]: Column names, or an empty list if the file cannot be read.
    """
    source = None
    try:
        source = build_row_source(path, settings)
        return list(source.columns)
    except Exception:
        _itk_log.exception("Could not read columns of %s", path)
        return []
    finally:
        if source is not None:
            source.close()


def format_cell(value) -> str:
    """Return a compact display string for one cell value.

    Args:
        value: Raw cell value from the frame.

    Returns:
        str: Text for the preview, with floats trimmed of representation noise.
    """
    if value is None:
        return ""
    if isinstance(value, float):
        if np.isnan(value):
            return ""
        if value == int(value) and abs(value) < 1e15:
            return str(int(value))
        return f"{value:.6g}"
    return str(value)


class LazyPreviewModel(QAbstractTableModel):
    """Table model that reveals a file progressively as the view scrolls.

    Rows and columns are both windowed. Rows grow downwards as the view asks
    for more through Qt's own ``fetchMore`` mechanism; columns grow rightwards
    through :meth:`fetch_more_columns`, which the view calls when its
    horizontal scrollbar nears the end. Qt has no column equivalent of
    ``fetchMore``, so that half is driven explicitly.

    Windowing the columns matters for wide instrument exports: laying out and
    measuring several hundred columns, each with its own isotope badge, is slow
    enough to be felt when only the first handful are ever looked at.
    """

    rowsAppended = Signal(int)
    columnsRevealed = Signal(int)
    sourceExhausted = Signal(int)
    headerRowChanged = Signal(int)

    def __init__(self, source: RowSource, parent=None, header_row: int | None = None):
        """Prime the model with the first window of rows and columns.

        Args:
            source (RowSource): Reader supplying rows on demand.
            parent: Optional Qt parent.
            header_row (int | None): In raw mode, which row holds the column
                names. None means the source already applied its own header.
        """
        super().__init__(parent)
        self._source = source
        self._header_row = header_row
        self._columns = list(source.columns)
        self._frame = pd.DataFrame(columns=self._columns)
        self._visible = 0
        self._visible_columns = min(INITIAL_VISIBLE_COLUMNS, len(self._columns))
        self._excluded_columns: set[int] = set()
        self._excluded_rows: set[int] = set()
        self._tints: dict[int, QColor] = {}
        self._muted = QColor(140, 140, 140)
        self._accent = QColor(60, 130, 220)
        self._strike_font: QFont | None = None
        self._header_font_cache: QFont | None = None
        self._ensure_loaded(INITIAL_VISIBLE_ROWS)
        self._visible = min(INITIAL_VISIBLE_ROWS, len(self._frame))

    @property
    def columns(self) -> list[str]:
        """Return the column names, taken from the header row in raw mode."""
        if self._header_row is None:
            return list(self._columns)
        return self._labels_from_row(self._header_row)

    def header_row(self) -> int | None:
        """Return which row holds the column names, or None outside raw mode."""
        return self._header_row

    def is_raw(self) -> bool:
        """Return True when the model is showing the file line for line."""
        return self._header_row is not None

    def set_header_row(self, row: int) -> bool:
        """Declare one row to be the header and relabel the columns.

        No re-read is needed: every line is already loaded, so changing the
        header is only a question of which row supplies the names and which
        rows count as data.

        Args:
            row (int): Zero-based row to treat as the header.

        Returns:
            bool: True when the header moved.
        """
        if self._header_row is None:
            return False
        row = max(0, min(int(row), max(0, len(self._frame) - 1)))
        if row == self._header_row:
            return False
        self._header_row = row
        self._source.scan_from = row + 1
        self._emit_full_refresh()
        self.headerDataChanged.emit(
            Qt.Horizontal, 0, max(0, self._visible_columns - 1))
        self.headerDataChanged.emit(Qt.Vertical, 0, max(0, self._visible - 1))
        self.headerRowChanged.emit(row)
        return True

    def data_row_count(self) -> int:
        """Return how many loaded rows sit below the header."""
        if self._header_row is None:
            return len(self._frame)
        return max(0, len(self._frame) - self._header_row - 1)

    def to_data_row(self, row: int) -> int:
        """Convert a preview row number into a data row number.

        Args:
            row (int): Zero-based row as numbered in the preview.

        Returns:
            int: Zero-based row as the importer will number it.
        """
        if self._header_row is None:
            return row
        return row - self._header_row - 1

    def _labels_from_row(self, row: int) -> list[str]:
        """Return the column names carried by one row.

        Args:
            row (int): Row holding the header.

        Returns:
            list[str]: One label per column, falling back to a position name.
        """
        labels = []
        for position in range(len(self._columns)):
            value = ""
            if 0 <= row < len(self._frame):
                value = format_cell(self._frame.iat[row, position]).strip()
            labels.append(value or f"Column {position + 1}")
        return labels

    def total_column_count(self) -> int:
        """Return how many columns the file has, revealed or not."""
        return len(self._columns)

    def can_fetch_more_columns(self) -> bool:
        """Return True while the file has columns still to reveal."""
        return self._visible_columns < len(self._columns)

    def fetch_more_columns(self, step: int = FETCH_STEP_COLUMNS) -> bool:
        """Reveal the next block of columns.

        Args:
            step (int): How many further columns to expose.

        Returns:
            bool: True when columns were revealed, False when none were left.
        """
        if not self.can_fetch_more_columns():
            return False
        new_total = min(self._visible_columns + max(1, step), len(self._columns))
        before = self.columnCount()
        self._visible_columns = new_total
        after = self.columnCount()
        if after > before:
            self.beginInsertColumns(QModelIndex(), before, after - 1)
            self.endInsertColumns()
        else:
            self.headerDataChanged.emit(Qt.Horizontal, 0, after - 1)
            self._emit_full_refresh()
        self.columnsRevealed.emit(new_total)
        return True

    def reveal_all_columns(self) -> int:
        """Reveal every remaining column at once.

        Returns:
            int: The number of columns now visible.
        """
        if self.can_fetch_more_columns():
            self.fetch_more_columns(len(self._columns))
        return self._visible_columns

    def source_path(self) -> str:
        """Return the path of the file being previewed."""
        return self._source.path

    def loaded_row_count(self) -> int:
        """Return how many rows have been pulled from disk so far."""
        return len(self._frame)

    def is_exhausted(self) -> bool:
        """Return True when every row in the file has been loaded."""
        return self._source.is_exhausted()

    def truncated_at(self) -> int | None:
        """Return the row where trailing text ended the data, if it did."""
        return self._source.truncated_at()

    def frame(self) -> pd.DataFrame:
        """Return the rows loaded so far as a frame."""
        return self._frame

    def close(self) -> None:
        """Release the underlying reader."""
        self._source.close()

    def set_muted_colour(self, colour: QColor) -> None:
        """Set the foreground colour used for removed rows and columns.

        Args:
            colour (QColor): Colour for struck-through cells.
        """
        self._muted = QColor(colour)
        self._emit_full_refresh()

    def set_excluded_columns(self, indices: set[int]) -> None:
        """Mark which column positions are removed.

        Args:
            indices (set[int]): Column positions the user has removed.
        """
        if indices == self._excluded_columns:
            return
        self._excluded_columns = set(indices)
        self._emit_full_refresh()
        self.headerDataChanged.emit(
            Qt.Horizontal, 0, max(0, self._visible_columns - 1))

    def set_excluded_rows(self, rows: set[int]) -> None:
        """Mark which absolute row numbers are removed.

        Args:
            rows (set[int]): Zero-based row numbers the user has removed.
        """
        if rows == self._excluded_rows:
            return
        self._excluded_rows = set(rows)
        self._emit_full_refresh()
        self.headerDataChanged.emit(Qt.Vertical, 0, max(0, self._visible - 1))

    def set_column_tints(self, tints: dict[int, QColor]) -> None:
        """Replace the per-column background tints used to show mappings.

        Args:
            tints (dict[int, QColor]): Column position to background colour.
        """
        self._tints = {k: QColor(v) for k, v in tints.items()}
        self._emit_full_refresh()

    def rowCount(self, parent=QModelIndex()) -> int:
        """Return the number of rows the grid shows, padding to a minimum.

        A short file still draws a full grid. Without the padding a two-row
        file would render as two stripes floating in grey, which reads as a
        failed load rather than as a small file.

        Args:
            parent: Unused parent index required by the Qt interface.
        """
        if parent.isValid():
            return 0
        return max(self._visible, MIN_GRID_ROWS)

    def columnCount(self, parent=QModelIndex()) -> int:
        """Return the number of columns the grid shows, padding to a minimum.

        Args:
            parent: Unused parent index required by the Qt interface.
        """
        if parent.isValid():
            return 0
        return max(self._visible_columns, MIN_GRID_COLUMNS)

    def real_row_count(self) -> int:
        """Return how many rows actually hold file content."""
        return self._visible

    def real_column_count(self) -> int:
        """Return how many columns actually exist in the file."""
        return self._visible_columns

    def canFetchMore(self, parent=QModelIndex()) -> bool:
        """Return True while more rows can be revealed or read from disk.

        Args:
            parent: Unused parent index required by the Qt interface.
        """
        if parent.isValid():
            return False
        if self._visible < len(self._frame):
            return True
        return not self._source.is_exhausted()

    def fetchMore(self, parent=QModelIndex()) -> None:
        """Reveal the next page of rows, reading from disk if needed.

        Args:
            parent: Unused parent index required by the Qt interface.
        """
        if parent.isValid():
            return
        target = self._visible + FETCH_STEP_ROWS
        self._ensure_loaded(target)
        new_visible = min(target, len(self._frame))
        if new_visible <= self._visible:
            if self._source.is_exhausted():
                self.sourceExhausted.emit(len(self._frame))
            return
        before = self.rowCount()
        self._visible = new_visible
        after = self.rowCount()
        if after > before:
            self.beginInsertRows(QModelIndex(), before, after - 1)
            self.endInsertRows()
        else:
            self._emit_full_refresh()
        self.rowsAppended.emit(self._visible)
        if self._source.is_exhausted() and self._visible >= len(self._frame):
            self.sourceExhausted.emit(len(self._frame))

    def data(self, index: QModelIndex, role: int = Qt.DisplayRole):
        """Return the value or styling for one cell.

        Args:
            index (QModelIndex): Cell being painted.
            role (int): Qt item role being requested.
        """
        if not index.isValid():
            return None
        row, col = index.row(), index.column()
        if row >= len(self._frame) or col >= self._visible_columns:
            if role == Qt.BackgroundRole:
                return QBrush(self._padding_tint())
            return None

        removed = col in self._excluded_columns or row in self._excluded_rows
        is_header = self._header_row is not None and row == self._header_row
        above_header = self._header_row is not None and row < self._header_row

        if role == Qt.DisplayRole:
            return format_cell(self._frame.iat[row, col])
        if role == Qt.FontRole:
            if removed:
                return self._strikeout_font()
            if is_header:
                return self._header_font()
            return None
        if role == Qt.ForegroundRole:
            if removed or above_header:
                return QBrush(self._muted)
            return None
        if role == Qt.BackgroundRole:
            if is_header:
                return QBrush(self._header_tint())
            if not removed and col in self._tints:
                return QBrush(self._tints[col])
            return None
        if role == Qt.TextAlignmentRole:
            value = self._frame.iat[row, col]
            if isinstance(value, (int, float, np.integer, np.floating)):
                return int(Qt.AlignRight | Qt.AlignVCenter)
            return int(Qt.AlignLeft | Qt.AlignVCenter)
        if role == Qt.ToolTipRole:
            if removed:
                return "Removed from import"
            if is_header:
                return "These are the column names"
            if above_header:
                return "Above the header, so not imported"
            return None
        return None

    def _header_font(self) -> QFont:
        """Return the cached bold font used for the header row."""
        if self._header_font_cache is None:
            font = QFont()
            font.setBold(True)
            self._header_font_cache = font
        return self._header_font_cache

    def _header_tint(self) -> QColor:
        """Return the background wash marking the header row."""
        tint = QColor(self._accent)
        tint.setAlpha(70)
        return tint

    def _padding_tint(self) -> QColor:
        """Return the faint wash used for grid cells the file does not reach."""
        tint = QColor(self._muted)
        tint.setAlpha(18)
        return tint

    def set_accent_colour(self, colour: QColor) -> None:
        """Set the colour used to mark the header row.

        Args:
            colour (QColor): Accent colour from the active palette.
        """
        self._accent = QColor(colour)
        self._emit_full_refresh()

    def headerData(self, section: int, orientation: Qt.Orientation,
                   role: int = Qt.DisplayRole):
        """Return the label or styling for one header section.

        Args:
            section (int): Row or column position.
            orientation (Qt.Orientation): Which header is being painted.
            role (int): Qt item role being requested.
        """
        if orientation == Qt.Horizontal:
            if section >= self._visible_columns:
                return "" if role == Qt.DisplayRole else None
            labels = self.columns
            name = labels[section] if section < len(labels) else ""
            if role == Qt.DisplayRole:
                return name
            if role == Qt.FontRole and section in self._excluded_columns:
                return self._strikeout_font()
            if role == Qt.ForegroundRole and section in self._excluded_columns:
                return QBrush(self._muted)
            if role == Qt.ToolTipRole:
                state = ("removed from import"
                         if section in self._excluded_columns else "included")
                return f"{name} — {state}"
            return None

        if section >= self._visible:
            return "" if role == Qt.DisplayRole else None
        if role == Qt.DisplayRole:
            if self._header_row is not None:
                if section == self._header_row:
                    return "▸"
                if section < self._header_row:
                    return "·"
                return str(section - self._header_row)
            return str(section + 1)
        if role == Qt.FontRole:
            if section in self._excluded_rows:
                return self._strikeout_font()
            if self._header_row is not None and section == self._header_row:
                return self._header_font()
            return None
        if role == Qt.ForegroundRole:
            if section in self._excluded_rows:
                return QBrush(self._muted)
            if self._header_row is not None and section < self._header_row:
                return QBrush(self._muted)
            return None
        if role == Qt.ToolTipRole:
            if self._header_row is None:
                return None
            if section == self._header_row:
                return "Header row — these cells name the columns"
            if section < self._header_row:
                return ("Above the header, so not imported. "
                        "Right-click to make this the header row.")
            return f"Data row {section - self._header_row}"
        return None

    def flags(self, index: QModelIndex):
        """Return the interaction flags for one cell.

        Args:
            index (QModelIndex): Cell being queried.
        """
        if not index.isValid():
            return Qt.NoItemFlags
        if (index.row() >= self._visible
                or index.column() >= self._visible_columns):
            return Qt.ItemIsEnabled
        return Qt.ItemIsEnabled | Qt.ItemIsSelectable

    def load_until(self, row_count: int) -> int:
        """Reveal rows up to ``row_count``, reading from disk as needed.

        Args:
            row_count (int): Row count the caller wants visible.

        Returns:
            int: The number of rows actually visible afterwards.
        """
        self._ensure_loaded(row_count)
        new_visible = min(row_count, len(self._frame))
        if new_visible > self._visible:
            before = self.rowCount()
            self._visible = new_visible
            after = self.rowCount()
            if after > before:
                self.beginInsertRows(QModelIndex(), before, after - 1)
                self.endInsertRows()
            else:
                self._emit_full_refresh()
            self.rowsAppended.emit(self._visible)
        return self._visible

    def load_all(self, hard_limit: int = 5_000_000) -> int:
        """Read the whole file into the model, up to a safety ceiling.

        Args:
            hard_limit (int): Maximum rows to load before stopping.

        Returns:
            int: The number of rows visible afterwards.
        """
        while not self._source.is_exhausted() and len(self._frame) < hard_limit:
            before = len(self._frame)
            self._ensure_loaded(before + DISK_CHUNK_ROWS)
            if len(self._frame) == before:
                break
        return self.load_until(len(self._frame))

    def _ensure_loaded(self, row_count: int) -> None:
        """Read from the source until ``row_count`` rows are buffered.

        Args:
            row_count (int): Number of rows the model needs available.
        """
        while len(self._frame) < row_count and not self._source.is_exhausted():
            needed = row_count - len(self._frame)
            chunk = self._source.fetch(max(needed, FETCH_STEP_ROWS))
            if chunk is None or not len(chunk):
                break
            if self._frame.empty:
                self._frame = chunk.reset_index(drop=True)
            else:
                self._frame = pd.concat(
                    [self._frame, chunk], ignore_index=True)

    def _strikeout_font(self) -> QFont:
        """Return the cached strikeout font used for removed cells."""
        if self._strike_font is None:
            font = QFont()
            font.setStrikeOut(True)
            self._strike_font = font
        return self._strike_font

    def _emit_full_refresh(self) -> None:
        """Repaint every revealed cell after a styling change."""
        if self._visible == 0 or not self._visible_columns:
            return
        top = self.index(0, 0)
        bottom = self.index(self._visible - 1, self._visible_columns - 1)
        self.dataChanged.emit(
            top, bottom,
            [Qt.ForegroundRole, Qt.FontRole, Qt.BackgroundRole, Qt.ToolTipRole])
